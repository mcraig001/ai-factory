"""
Cash Flow Forecaster - REBUILT
Simple, clean, professional - no bullshit
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

class CashFlowForecaster:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Professional colors - blues and grays only
        self.colors = {
            'header': '1E40AF',        # Deep blue
            'subheader': '3B82F6',     # Medium blue
            'positive': '10B981',      # Green
            'negative': 'EF4444',      # Red
            'warning': 'F59E0B',       # Orange
            'input': 'FEF3C7',         # Light yellow
            'light_bg': 'F3F4F6',      # Light gray
        }
        
    def create_inputs(self):
        """Simple inputs sheet"""
        ws = self.wb.create_sheet("Inputs", 0)
        
        # Title
        ws['A1'] = 'BUSINESS INPUTS'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:C1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = 'Enter your data in the YELLOW cells below'
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:C2')
        
        # Inputs
        row = 4
        inputs = [
            ('Starting Cash', 50000, '$'),
            ('', '', ''),
            ('MONTHLY REVENUE', '', ''),
            ('Current Monthly Revenue', 25000, '$'),
            ('Revenue Growth % (monthly)', 0.05, '%'),
            ('', '', ''),
            ('MONTHLY EXPENSES', '', ''),
            ('Salary & Payroll', 12000, '$'),
            ('Rent & Facilities', 3000, '$'),
            ('Marketing & Sales', 2500, '$'),
            ('Software & Tools', 800, '$'),
            ('Other Operating Expenses', 1700, '$'),
        ]
        
        ws['A3'] = 'Description'
        ws['B3'] = 'Amount'
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        for desc, value, fmt in inputs:
            if desc == '':
                row += 1
                continue
                
            ws[f'A{row}'] = desc
            
            if value != '':
                ws[f'B{row}'] = value
                ws[f'B{row}'].fill = PatternFill(start_color=self.colors['input'], fill_type='solid')
                ws[f'B{row}'].font = Font(bold=True)
                
                if fmt == '$':
                    ws[f'B{row}'].number_format = '$#,##0'
                elif fmt == '%':
                    ws[f'B{row}'].number_format = '0.0%'
            else:
                ws[f'A{row}'].font = Font(bold=True, color='1E40AF')
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        
        print("  ✓ Inputs created")
        
    def create_projection(self):
        """12-month projection - ONE clean table"""
        ws = self.wb.create_sheet("12-Month Forecast", 1)
        
        # Title
        ws['A1'] = '12-MONTH CASH FLOW FORECAST'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:N1')
        ws.row_dimensions[1].height = 30
        
        # Month headers
        row = 3
        ws['A3'] = ''
        ws['B3'] = 'Starting'
        
        start_date = datetime.now().replace(day=1)
        for i in range(12):
            month_date = start_date + timedelta(days=30*i)
            col = get_column_letter(i + 3)
            ws[f'{col}{row}'] = month_date.strftime('%b %y')
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['subheader'], fill_type='solid')
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        # REVENUE
        row = 5
        ws[f'A{row}'] = 'REVENUE'
        ws[f'A{row}'].font = Font(size=12, bold=True, color='10B981')
        
        row += 1
        ws[f'A{row}'] = 'Monthly Revenue'
        for i in range(12):
            col = get_column_letter(i + 3)
            if i == 0:
                ws[f'{col}{row}'] = '=Inputs!$B$5'
            else:
                prev_col = get_column_letter(i + 2)
                ws[f'{col}{row}'] = f'={prev_col}{row}*(1+Inputs!$B$6)'
            ws[f'{col}{row}'].number_format = '$#,##0'
        
        row += 1
        ws[f'A{row}'] = 'Total Revenue'
        ws[f'A{row}'].font = Font(bold=True)
        for i in range(12):
            col = get_column_letter(i + 3)
            ws[f'{col}{row}'] = f'={col}{row-1}'
            ws[f'{col}{row}'].number_format = '$#,##0'
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light_bg'], fill_type='solid')
        
        # EXPENSES
        row += 2
        ws[f'A{row}'] = 'EXPENSES'
        ws[f'A{row}'].font = Font(size=12, bold=True, color='EF4444')
        
        expense_rows = [
            ('Salary & Payroll', 'Inputs!$B$9'),
            ('Rent & Facilities', 'Inputs!$B$10'),
            ('Marketing & Sales', 'Inputs!$B$11'),
            ('Software & Tools', 'Inputs!$B$12'),
            ('Other Operating', 'Inputs!$B$13'),
        ]
        
        for expense_name, cell_ref in expense_rows:
            row += 1
            ws[f'A{row}'] = expense_name
            for i in range(12):
                col = get_column_letter(i + 3)
                ws[f'{col}{row}'] = f'={cell_ref}'
                ws[f'{col}{row}'].number_format = '$#,##0'
        
        row += 1
        ws[f'A{row}'] = 'Total Expenses'
        ws[f'A{row}'].font = Font(bold=True)
        for i in range(12):
            col = get_column_letter(i + 3)
            # Sum all expense rows
            start_row = row - 5
            end_row = row - 1
            ws[f'{col}{row}'] = f'=SUM({col}{start_row}:{col}{end_row})'
            ws[f'{col}{row}'].number_format = '$#,##0'
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light_bg'], fill_type='solid')
        
        # NET CASH FLOW
        row += 2
        ws[f'A{row}'] = 'NET CASH FLOW'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        revenue_row = 7
        expenses_row = row - 1
        
        for i in range(12):
            col = get_column_letter(i + 3)
            ws[f'{col}{row}'] = f'={col}{revenue_row}-{col}{expenses_row}'
            ws[f'{col}{row}'].number_format = '$#,##0'
            ws[f'{col}{row}'].font = Font(bold=True, size=11)
        
        # ENDING CASH BALANCE
        row += 1
        ws[f'A{row}'] = 'ENDING CASH BALANCE'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        ws[f'B{row}'] = '=Inputs!$B$4'
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'B{row}'].font = Font(bold=True, size=11)
        
        net_flow_row = row - 1
        for i in range(12):
            col = get_column_letter(i + 3)
            prev_col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = f'={prev_col}{row}+{col}{net_flow_row}'
            ws[f'{col}{row}'].number_format = '$#,##0'
            ws[f'{col}{row}'].font = Font(bold=True, size=11)
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light_bg'], fill_type='solid')
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 12
        for i in range(12):
            col = get_column_letter(i + 3)
            ws.column_dimensions[col].width = 11
        
        print("  ✓ 12-Month Forecast created")
        
    def create_dashboard(self):
        """Simple dashboard summary"""
        ws = self.wb.create_sheet("Dashboard", 2)
        
        # Title
        ws['A1'] = 'CASH FLOW DASHBOARD'
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 35
        
        ws['A2'] = 'Executive Summary'
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:D2')
        
        # Key metrics
        row = 4
        ws['A4'] = 'Metric'
        ws['B4'] = 'Value'
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(bold=True)
        
        row = 5
        metrics = [
            ('Starting Cash', '=Inputs!B4'),
            ('Total Revenue (12 months)', '=SUM(\'12-Month Forecast\'!C7:N7)'),
            ('Total Expenses (12 months)', '=SUM(\'12-Month Forecast\'!C14:N14)'),
            ('Net Cash Flow (12 months)', '=B6-B7'),
            ('Ending Cash Balance', '=\'12-Month Forecast\'!N18'),
            ('Lowest Cash Point', '=MIN(\'12-Month Forecast\'!C18:N18)'),
        ]
        
        for metric, formula in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = formula
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'B{row}'].font = Font(bold=True, size=11)
            row += 1
        
        # Status check
        row += 2
        ws[f'A{row}'] = 'CASH HEALTH CHECK'
        ws[f'A{row}'].font = Font(size=12, bold=True, color='1E40AF')
        
        row += 1
        ws[f'A{row}'] = 'Status:'
        ws[f'B{row}'] = '=IF(B9>B4*0.5,"✅ Healthy","⚠️ Monitor Closely")'
        ws[f'B{row}'].font = Font(bold=True, size=11)
        
        row += 1
        ws[f'A{row}'] = 'Runway (months):'
        ws[f'B{row}'] = '=B9/(B7/12)'
        ws[f'B{row}'].number_format = '0.0'
        ws[f'B{row}'].font = Font(bold=True, size=11)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        print("  ✓ Dashboard created")
        
    def create_instructions(self):
        """Simple instructions"""
        ws = self.wb.create_sheet("How to Use", 3)
        
        ws['A1'] = 'HOW TO USE THIS FORECASTER'
        ws['A1'].font = Font(size=16, bold=True, color='1E40AF')
        
        instructions = [
            '',
            'STEP 1: Enter Your Data',
            '→ Go to "Inputs" sheet',
            '→ Fill in YELLOW cells with your numbers',
            '→ Starting cash, revenue, expenses',
            '',
            'STEP 2: Review Forecast',
            '→ Go to "12-Month Forecast" sheet',
            '→ See your cash position for next 12 months',
            '→ Check ending cash balance',
            '',
            'STEP 3: Check Dashboard',
            '→ Go to "Dashboard" sheet',
            '→ See summary metrics',
            '→ Check cash health status',
            '',
            'KEY INSIGHTS:',
            '→ Ending Cash Balance: How much cash you\'ll have in 12 months',
            '→ Lowest Cash Point: Your minimum cash position (watch this!)',
            '→ Runway: How many months until you run out (if growth stops)',
            '',
            'TIPS:',
            '• Update monthly with actual numbers',
            '• Be conservative with revenue growth',
            '• Include all expenses',
            '• Keep 3-6 months cash buffer',
            '',
            'Questions? Email: support@yoursite.com',
        ]
        
        row = 3
        for text in instructions:
            ws[f'A{row}'] = text
            if text.startswith('STEP') or text.startswith('KEY'):
                ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
        
        ws.column_dimensions['A'].width = 70
        
        print("  ✓ Instructions created")
        
    def build(self):
        """Build the complete forecaster"""
        print("\n🏗️  Building Cash Flow Forecaster (CLEAN VERSION)")
        print("="*60)
        
        self.create_inputs()
        self.create_projection()
        self.create_dashboard()
        self.create_instructions()
        
        # Save
        output_dir = "/home/claude/ai-factory/products/cash_flow_forecaster_pro/built"
        os.makedirs(output_dir, exist_ok=True)
        
        output_path = f"{output_dir}/cash-flow-forecaster-v2.xlsx"
        self.wb.save(output_path)
        
        print("\n" + "="*60)
        print("✅ CLEAN VERSION COMPLETE")
        print("="*60)
        print(f"File: {output_path}")
        print(f"Sheets: 4 (Inputs, 12-Month Forecast, Dashboard, How to Use)")
        print(f"Structure: SIMPLE and CLEAN")
        print(f"Ready to review")
        print("="*60 + "\n")
        
        return output_path

if __name__ == "__main__":
    forecaster = CashFlowForecaster()
    forecaster.build()
